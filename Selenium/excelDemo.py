import openpyxl

book=openpyxl.load_workbook("C:\\Jmeter\\Book1.xlsx")
sheet1=book.active
Fname=sheet1.cell(row=1,column=2)
print(Fname.value)
Dict={}
#sheet1.cell(row=2,column=2).value="Rahul"
print(sheet1.cell(row=2,column=2).value)
print(sheet1.max_row)
print(sheet1.max_column)
print(sheet1['A3'].value)

for i in range(1,sheet1.max_row+1):
    if sheet1.cell(row=i,column=1).value=="Testcase1":
        for j in range(2,sheet1.max_column+1):
            Dict[sheet1.cell(row=1,column=j).value]=sheet1.cell(row=i,column=j).value

print(Dict)
